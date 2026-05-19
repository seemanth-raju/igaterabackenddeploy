"""
iGatera Migration Extractor
============================
Run on the same local network as the Matrix device.

Two modes:
  Device Scan   — connect to any Matrix device (Panel Lite, ARGO, ARGO FACE, etc.),
                  scan all users, extract fingerprint templates, upload to iGatera.
  Excel Import  — upload a Matrix User_Configuration.xls export,
                  supply device credentials so the backend can extract face/finger
                  templates; card/PIN are read directly from the Excel.

Setup:
    pip install -r requirements.txt
    streamlit run app.py
"""

import io
import re
import xml.etree.ElementTree as ET

import openpyxl
import requests
import streamlit as st
import urllib3
from requests.auth import HTTPDigestAuth

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DEFAULT_BACKEND = "https://app.igatera.com:8099/api"

# ---------------------------------------------------------------------------
# Device client helpers  (shared by both modes)
# ---------------------------------------------------------------------------

class DeviceError(Exception):
    def __init__(self, message, *, kind="request", status_code=None, response_text=""):
        super().__init__(message)
        self.kind = kind
        self.status_code = status_code
        self.response_text = response_text


def _make_session(username: str, password: str) -> requests.Session:
    s = requests.Session()
    s.auth = HTTPDigestAuth(username, password)
    s.verify = False
    return s


def _request(session: requests.Session, method: str, url: str, *,
             params=None, data=None, timeout=(5, 20)):
    try:
        resp = getattr(session, method)(url, params=params, data=data, timeout=timeout)
    except requests.exceptions.ConnectionError as e:
        raise DeviceError("Cannot reach device. Check IP, port, and network path.", kind="connection") from e
    except requests.exceptions.Timeout as e:
        raise DeviceError("Request timed out.", kind="timeout") from e
    except requests.exceptions.RequestException as e:
        raise DeviceError(f"Request failed: {e}") from e
    if resp.status_code in (401, 403):
        raise DeviceError("Authentication failed — check username and password.", kind="auth",
                          status_code=resp.status_code, response_text=resp.text)
    if resp.status_code >= 400:
        raise DeviceError(f"Device returned HTTP {resp.status_code}.", kind="http",
                          status_code=resp.status_code, response_text=resp.text)
    return resp


def _base(ip: str, port: int) -> str:
    return f"http://{ip}:{port}/device.cgi"


def device_test_connection(ip: str, port: int, username: str, password: str) -> dict:
    session = _make_session(username, password)
    try:
        resp = _request(session, "get", f"{_base(ip, port)}/device-basic-config",
                        params={"action": "get", "format": "xml"}, timeout=(3, 6))
        root = ET.fromstring(resp.text)
        def _f(*tags):
            for t in tags:
                v = root.findtext(t)
                if v and v.strip():
                    return v.strip()
            return None
        return {"ok": True, "mac": _f("mac-address", "Mac-Address"),
                "model": _f("device-model", "Device-Model"),
                "serial": _f("serial-number", "Serial-Number")}
    except (DeviceError, ET.ParseError):
        pass
    _request(session, "get", f"{_base(ip, port)}/users",
             params={"action": "get", "user-id": "1", "format": "xml"}, timeout=(3, 6))
    return {"ok": True, "mac": None, "model": None, "serial": None}


def _parse_user(body: str) -> dict | None:
    body = (body or "").strip()
    if not body:
        return None
    if any(x in body for x in ("Response-Code=10", "<Response-Code>10</Response-Code>",
                                "Response-Code=13", "<Response-Code>13</Response-Code>",
                                "Request Failed")):
        return None
    try:
        root = ET.fromstring(body)
        rc = root.findtext("Response-Code") or root.findtext("response-code")
        if rc and rc.strip() not in ("0", ""):
            return None
        uid = (root.findtext("user-id") or "").strip()
        if uid:
            return {"user_id": uid,
                    "full_name": (root.findtext("name") or uid).strip(),
                    "is_active": (root.findtext("user-active") or "1").strip() != "0"}
    except ET.ParseError:
        pass
    def _e(tag):
        m = re.search(rf"<{tag}>(.*?)</{tag}>", body, re.IGNORECASE | re.DOTALL)
        return m.group(1).strip() if m else ""
    uid = _e("user-id")
    if not uid:
        return None
    return {"user_id": uid, "full_name": _e("name") or uid,
            "is_active": (_e("user-active") or "1").strip() != "0"}


def device_get_user(ip: str, port: int, username: str, password: str, user_id: str) -> dict | None:
    session = _make_session(username, password)
    resp = _request(session, "get", f"{_base(ip, port)}/users",
                    params={"action": "get", "user-id": user_id, "format": "xml"}, timeout=(5, 10))
    return _parse_user(resp.text)


def device_get_fingerprint(ip: str, port: int, username: str, password: str,
                           user_id: str, finger_index: int) -> bytes | None:
    session = _make_session(username, password)
    params = {"action": "get", "type": "1", "user-id": user_id, "finger-index": finger_index}
    for method in ("post", "get"):
        try:
            resp = _request(session, method, f"{_base(ip, port)}/credential",
                            params=params, timeout=(5, 30))
            if not resp.content:
                continue
            c = resp.content.strip()
            if (b"Request Failed" in c or b"Response-Code=" in c
                    or (c.startswith(b"<") and b"Response-Code" in c)):
                continue
            return resp.content
        except DeviceError as e:
            if e.kind in ("auth", "connection", "timeout"):
                raise
            continue
    return None


def extract_fingerprints_for_users(ip: str, port: int, username: str, password: str,
                                   user_ids: list[str], progress_placeholder=None) -> dict:
    results: dict[str, dict[int, bytes]] = {}
    total = len(user_ids)
    for i, uid in enumerate(user_ids):
        user_fps: dict[int, bytes] = {}
        consecutive_empty = 0
        for finger_idx in range(1, 11):
            data = device_get_fingerprint(ip, port, username, password, uid, finger_idx)
            if data:
                user_fps[finger_idx] = data
                consecutive_empty = 0
            else:
                consecutive_empty += 1
                if user_fps and consecutive_empty >= 3:
                    break
        if user_fps:
            results[uid] = user_fps
        if progress_placeholder is not None:
            pct = (i + 1) / max(total, 1)
            fp_count = sum(len(v) for v in results.values())
            progress_placeholder.progress(pct,
                text=f"Users checked: {i + 1}/{total} — {fp_count} template(s) found")
    return results


# ---------------------------------------------------------------------------
# Backend helpers
# ---------------------------------------------------------------------------

def _backend_headers(token: str | None) -> dict:
    if token and token.strip():
        return {"Authorization": f"Bearer {token.strip()}"}
    return {}


def fetch_groups(backend_api: str, token: str | None) -> list[dict]:
    """Return list of {group_id, name, code} from backend, or [] on failure."""
    try:
        resp = requests.get(f"{backend_api}/groups", headers=_backend_headers(token),
                            timeout=(5, 15), verify=False)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return []


def _parse_excel_preview(excel_bytes: bytes) -> dict:
    """Return a quick summary of the Excel without full parsing."""
    import io as _io
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(excel_bytes), read_only=True, data_only=True)
        ws = wb.active
        headers = []
        rows = 0
        card_count = face_count = pin_count = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(c).strip().lower().replace(" ", "_") if c else "" for c in row]
                continue
            if not any(row):
                continue
            row_dict = {headers[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row) if i < len(headers) and headers[i]}
            rows += 1
            if row_dict.get("card1", "").strip():
                card_count += 1
            if row_dict.get("pin", "").strip():
                pin_count += 1
            if row_dict.get("face_recognition", "0").strip() == "1":
                face_count += 1
        wb.close()
        return {"rows": rows, "card_count": card_count,
                "pin_count": pin_count, "face_count": face_count,
                "has_matrix_format": "user_name" in headers or "face_recognition" in headers}
    except Exception as e:
        return {"error": str(e)}


# ---------------------------------------------------------------------------
# Streamlit page setup
# ---------------------------------------------------------------------------

st.set_page_config(page_title="iGatera Migration Extractor", page_icon="🔄", layout="wide")
st.title("🔄 iGatera Migration Extractor")

# ---------------------------------------------------------------------------
# Sidebar — shared settings
# ---------------------------------------------------------------------------
with st.sidebar:
    st.header("☁️ iGatera Backend")
    backend_api = st.text_input("Backend API URL", value=DEFAULT_BACKEND,
                                help="e.g. https://app.igatera.com:8099/api")
    auth_token = st.text_input("Auth Token (optional)",
                               type="password",
                               help="Bearer token — enables group lookup from backend")

    st.divider()
    st.header("📋 Import Settings")

    # Group selection — fetch from backend if token provided
    groups = []
    if auth_token.strip():
        if st.button("🔄 Fetch Groups", use_container_width=True):
            with st.spinner("Fetching groups..."):
                groups = fetch_groups(backend_api, auth_token)
            st.session_state["groups"] = groups
        groups = st.session_state.get("groups", [])

    if groups:
        group_options = {f"{g['name']} (ID: {g['group_id']})": g["group_id"] for g in groups}
        selected_group_label = st.selectbox("Group", list(group_options.keys()))
        group_id = group_options[selected_group_label]
    else:
        group_id = st.number_input("Group ID", min_value=1, value=1, step=1,
                                   help="Provide auth token above to load groups from backend")

    site_id = st.number_input("Site ID", min_value=1, value=1, step=1)
    company_id_override = st.text_input("Company ID (super-admin only)", value="")

    st.divider()
    st.caption(f"Backend: `{backend_api}`")

# ---------------------------------------------------------------------------
# Mode tabs
# ---------------------------------------------------------------------------
tab_device, tab_excel = st.tabs(["🖧 Device Scan", "📊 Excel Import"])

# ===========================================================================
# TAB 1 — Device Scan  (Panel Lite / ARGO / any Matrix device)
# ===========================================================================
with tab_device:
    st.subheader("Device Scan Migration")
    st.caption(
        "Connect to any Matrix device (Panel Lite, ARGO, ARGO FACE, etc.), "
        "fetch users by ID, extract fingerprint templates, then upload to iGatera."
    )

    # Session state for this tab
    for key, default in {
        "dev_connected": False, "dev_info": None,
        "dev_profiles": None, "dev_fingerprints": None,
    }.items():
        if key not in st.session_state:
            st.session_state[key] = default

    # ── Device connection ──────────────────────────────────────────────────
    with st.expander("Step 1 — Connect to Device", expanded=not st.session_state["dev_connected"]):
        col1, col2 = st.columns(2)
        with col1:
            dev_ip   = st.text_input("Device IP", value="192.168.1.100", key="dev_ip")
            dev_user = st.text_input("Username", value="admin", key="dev_user")
        with col2:
            dev_port = st.number_input("Port", value=1040, min_value=1, max_value=65535,
                                       step=1, key="dev_port",
                                       help="Panel Lite: 1040  |  ARGO / ARGO FACE: 80")
            dev_pass = st.text_input("Password", type="password", key="dev_pass")

        dev_vendor = st.text_input("Device Vendor", value="Matrix", key="dev_vendor")
        dev_model  = st.text_input("Device Model", value="Panel Lite", key="dev_model",
                                   help="e.g. Panel Lite, COSEC ARGO, COSEC ARGO FACE")

        if st.button("🔌 Test Connection", use_container_width=True):
            if not dev_ip or not dev_pass:
                st.error("Enter IP and password.")
            else:
                try:
                    with st.spinner("Connecting..."):
                        info = device_test_connection(dev_ip, dev_port, dev_user, dev_pass)
                    st.session_state["dev_connected"] = True
                    st.session_state["dev_info"] = info
                    parts = ["✅ Connected"]
                    if info.get("model"):
                        parts.append(f"Model: {info['model']}")
                    if info.get("mac"):
                        parts.append(f"MAC: {info['mac']}")
                    st.success(" | ".join(parts))
                except DeviceError as e:
                    st.session_state["dev_connected"] = False
                    st.error(str(e))

    if st.session_state["dev_connected"]:
        info = st.session_state["dev_info"] or {}
        c1, c2, c3 = st.columns(3)
        c1.success("Connected")
        if info.get("model"):
            c2.info(f"Model: {info['model']}")
        if info.get("mac"):
            c3.info(f"MAC: {info['mac']}")

    # ── Fetch users ────────────────────────────────────────────────────────
    st.divider()
    with st.expander("Step 2 — Fetch Users by ID", expanded=True):
        if not st.session_state["dev_connected"]:
            st.info("Connect to the device first (Step 1).")
        else:
            st.caption("Enter one or more User IDs (comma-separated). Example: `2002001, 2002002`")
            user_ids_input = st.text_area("User ID(s)", height=80,
                                          placeholder="2002001, 2002002, 2002003")

            if st.button("🔎 Fetch User(s)", type="primary", use_container_width=True):
                raw_ids = [x.strip() for x in user_ids_input.replace("\n", ",").split(",") if x.strip()]
                if not raw_ids:
                    st.error("Enter at least one User ID.")
                else:
                    found, not_found, errors = [], [], []
                    prog = st.progress(0.0, text="Fetching users...")
                    for i, uid in enumerate(raw_ids):
                        try:
                            profile = device_get_user(dev_ip, dev_port, dev_user, dev_pass, uid)
                            (found if profile else not_found).append(profile or uid)
                        except DeviceError as e:
                            errors.append(f"{uid}: {e}")
                        prog.progress((i + 1) / len(raw_ids), text=f"Checked {i + 1}/{len(raw_ids)}")
                    prog.progress(1.0, text="Done")
                    st.session_state["dev_profiles"] = found or None
                    if found:
                        st.success(f"Found **{len(found)}** user(s).")
                    if not_found:
                        st.warning(f"Not found: {', '.join(str(x) for x in not_found)}")
                    for err in errors:
                        st.error(err)

    if st.session_state.get("dev_profiles"):
        fps_map = {uid: len(fps) for uid, fps in (st.session_state.get("dev_fingerprints") or {}).items()}
        st.dataframe(
            [{"User ID": p["user_id"], "Full Name": p["full_name"],
              "Active": "Yes" if p["is_active"] else "No",
              "Fingerprints": fps_map.get(p["user_id"], 0)}
             for p in st.session_state["dev_profiles"]],
            use_container_width=True, hide_index=True,
        )

    # ── Extract fingerprints ───────────────────────────────────────────────
    st.divider()
    with st.expander("Step 3 — Extract Fingerprints"):
        if not st.session_state.get("dev_profiles"):
            st.info("Fetch users first (Step 2).")
        else:
            if st.button("🖐 Extract Fingerprints", type="primary", use_container_width=True):
                user_ids = [p["user_id"] for p in st.session_state["dev_profiles"]]
                fp_prog = st.progress(0.0, text="Starting...")
                try:
                    fps = extract_fingerprints_for_users(
                        dev_ip, dev_port, dev_user, dev_pass, user_ids,
                        progress_placeholder=fp_prog)
                    st.session_state["dev_fingerprints"] = fps
                    total_fp = sum(len(v) for v in fps.values())
                    fp_prog.progress(1.0, text=f"Done — {total_fp} template(s)")
                    if total_fp == 0:
                        st.warning("No fingerprint templates found.")
                    else:
                        st.success(f"Extracted **{total_fp}** template(s) from **{len(fps)}** user(s).")
                except DeviceError as e:
                    st.error(str(e))
                    st.session_state["dev_fingerprints"] = None

            fps = st.session_state.get("dev_fingerprints") or {}
            if fps:
                total_fp = sum(len(v) for v in fps.values())
                c1, c2, c3 = st.columns(3)
                c1.metric("Users with Fingerprints", len(fps))
                c2.metric("Total Templates", total_fp)
                c3.metric("Avg / User", f"{total_fp / max(len(fps), 1):.1f}")

    # ── Upload ─────────────────────────────────────────────────────────────
    st.divider()
    st.subheader("Step 4 — Upload to iGatera")

    if not st.session_state.get("dev_profiles"):
        st.info("Complete Steps 2 and 3 first.")
    else:
        fingerprints = st.session_state.get("dev_fingerprints") or {}
        total_fp = sum(len(v) for v in fingerprints.values())
        profiles   = st.session_state["dev_profiles"]
        device_info = st.session_state.get("dev_info") or {}

        c1, c2 = st.columns(2)
        c1.metric("Users to upload", len(profiles))
        c2.metric("Fingerprint templates", total_fp)

        if st.button("📤 Upload to iGatera", type="primary", use_container_width=True,
                     key="dev_upload"):
            with st.spinner("Building Excel..."):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Users"
                ws.append(["user_id", "ref_user_id", "full_name", "is_active", "valid_till", "user_index"])
                for p in profiles:
                    ws.append([p["user_id"], p["user_id"], p["full_name"],
                               "1" if p["is_active"] else "0", "", ""])
                buf = io.BytesIO()
                wb.save(buf)
                excel_bytes = buf.getvalue()

            files = [("users_excel",
                      ("users.xlsx", excel_bytes,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))]
            for uid, fp_dict in fingerprints.items():
                for finger_idx, fp_bytes in fp_dict.items():
                    files.append(("fingerprints",
                                  (f"{uid}_finger_{finger_idx}.dat", fp_bytes,
                                   "application/octet-stream")))

            form_data = {
                "group_id": str(int(group_id)),
                "site_id":  str(int(site_id)),
                "device_vendor": dev_vendor or "Matrix",
                "device_model":  dev_model or "Panel Lite",
            }
            if device_info.get("mac"):
                form_data["device_mac"] = device_info["mac"]
            if device_info.get("serial"):
                form_data["device_serial"] = device_info["serial"]
            if company_id_override.strip():
                form_data["company_id"] = company_id_override.strip()

            headers = _backend_headers(auth_token)

            with st.spinner(f"Uploading {len(profiles)} user(s) + {total_fp} template(s)..."):
                try:
                    resp = requests.post(f"{backend_api}/devices/upload-import",
                                         data=form_data, files=files, headers=headers,
                                         timeout=(15, 180), verify=False)
                    if resp.status_code == 201:
                        result = resp.json()
                        st.success("Upload complete!")
                        r1, r2, r3, r4 = st.columns(4)
                        r1.metric("Imported",   result.get("imported_user_count", 0))
                        r2.metric("Created",    result.get("created_tenants", 0))
                        r3.metric("Updated",    result.get("updated_tenants", 0))
                        r4.metric("Fingerprints", result.get("imported_fingerprint_count", 0))
                        for w in result.get("warnings", []):
                            st.warning(w)
                        with st.expander("Per-user results"):
                            st.dataframe(result.get("users", []), use_container_width=True)
                    else:
                        st.error(f"Upload failed: HTTP {resp.status_code}")
                        try:
                            st.code(str(resp.json().get("detail", resp.text[:500])))
                        except Exception:
                            st.code(resp.text[:500])
                except requests.exceptions.ConnectionError:
                    st.error(f"Cannot reach backend at `{backend_api}`.")
                except requests.exceptions.Timeout:
                    st.error("Timed out. Check the database before retrying to avoid duplicates.")
                except Exception as exc:
                    st.error(f"Unexpected error: {exc}")

# ===========================================================================
# TAB 2 — Excel Import  (Matrix User_Configuration.xls)
# ===========================================================================
with tab_excel:
    st.subheader("Excel Import Migration")
    st.caption(
        "Upload a **Matrix User_Configuration.xls** export. "
        "Card and PIN are read directly from the Excel. "
        "Face and finger templates are extracted from the device by the backend — "
        "supply device credentials below."
    )

    # Session state for this tab
    if "xl_preview" not in st.session_state:
        st.session_state["xl_preview"] = None

    # ── Excel upload ───────────────────────────────────────────────────────
    st.markdown("#### Step 1 — Upload Excel")
    uploaded_file = st.file_uploader(
        "Matrix User_Configuration.xls / .xlsx",
        type=["xls", "xlsx"],
        help="Exported from Matrix COSEC software: Reports → User Configuration",
    )

    if uploaded_file is not None:
        excel_bytes = uploaded_file.read()
        with st.spinner("Parsing Excel..."):
            preview = _parse_excel_preview(excel_bytes)
        st.session_state["xl_preview"] = {"bytes": excel_bytes, "stats": preview,
                                          "filename": uploaded_file.name}

    preview_data = st.session_state.get("xl_preview")
    if preview_data:
        stats = preview_data["stats"]
        if "error" in stats:
            st.error(f"Cannot parse Excel: {stats['error']}")
        else:
            if stats.get("has_matrix_format"):
                st.success(f"✅ Matrix format detected — **{stats['rows']}** users")
            else:
                st.info(f"Generic format — **{stats['rows']}** users (card/PIN/face columns not found)")

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total Users",    stats["rows"])
            c2.metric("Have Card",      stats["card_count"])
            c3.metric("Have Face",      stats["face_count"])
            c4.metric("Have PIN",       stats["pin_count"])

            if not stats.get("has_matrix_format"):
                st.warning(
                    "Matrix-format columns (User Name, Card1, Face Recognition, etc.) "
                    "were not detected. Card, PIN, and face extraction will be skipped. "
                    "Make sure the Excel is the full User_Configuration export."
                )

    # ── Device credentials ─────────────────────────────────────────────────
    st.divider()
    st.markdown("#### Step 2 — Device Credentials (for finger & face extraction)")
    st.caption("Optional — if not provided, only card/PIN from Excel will be imported.")

    col1, col2 = st.columns(2)
    with col1:
        xl_dev_ip   = st.text_input("Device IP Address", placeholder="192.168.1.x", key="xl_ip")
        xl_dev_user = st.text_input("Device Username", value="admin", key="xl_user")
    with col2:
        xl_dev_port = st.number_input("Device Port", value=80, min_value=1, max_value=65535,
                                      step=1, key="xl_port",
                                      help="ARGO / ARGO FACE: 80  |  Panel Lite: 1040")
        xl_dev_pass = st.text_input("Device Password", type="password", key="xl_pass")

    xl_dev_vendor = st.text_input("Device Vendor", value="Matrix", key="xl_vendor")
    xl_dev_model  = st.text_input("Device Model", placeholder="e.g. COSEC ARGO FACE",
                                  key="xl_model")

    # Optional connection test
    if st.button("🔌 Test Device Connection", key="xl_test_conn"):
        if not xl_dev_ip or not xl_dev_pass:
            st.warning("Enter IP and password to test the connection.")
        else:
            try:
                with st.spinner("Connecting..."):
                    info = device_test_connection(xl_dev_ip, xl_dev_port, xl_dev_user, xl_dev_pass)
                parts = ["✅ Connected"]
                if info.get("model"):
                    parts.append(f"Model: {info['model']}")
                if info.get("mac"):
                    parts.append(f"MAC: {info['mac']}")
                st.success(" | ".join(parts))
                st.session_state["xl_dev_info"] = info
            except DeviceError as e:
                st.error(str(e))
                st.session_state["xl_dev_info"] = None

    # ── Upload ─────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("#### Step 3 — Upload to iGatera")

    has_excel = preview_data and "error" not in preview_data.get("stats", {})
    if not has_excel:
        st.info("Upload an Excel file first (Step 1).")
    else:
        stats = preview_data["stats"]

        # Summary of what will be imported
        will_extract = []
        if stats["card_count"] > 0:
            will_extract.append(f"**{stats['card_count']}** card credential(s) from Excel")
        if stats["pin_count"] > 0:
            will_extract.append(f"**{stats['pin_count']}** PIN(s) from Excel")
        if xl_dev_ip and xl_dev_pass:
            if stats["face_count"] > 0:
                will_extract.append(f"**{stats['face_count']}** face template(s) from device")
            will_extract.append("finger templates from device (all users)")
        else:
            will_extract.append("⚠️ No device credentials — biometrics will be skipped")

        st.info("**What will be imported:**\n- " + "\n- ".join(will_extract))

        if st.button("📤 Upload to iGatera", type="primary", use_container_width=True,
                     key="xl_upload"):
            xl_dev_info = st.session_state.get("xl_dev_info") or {}
            excel_bytes = preview_data["bytes"]
            filename    = preview_data["filename"]

            files = [("users_excel",
                      (filename, excel_bytes,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                       if filename.endswith(".xlsx") else "application/vnd.ms-excel"))]

            form_data: dict = {
                "group_id": str(int(group_id)),
                "site_id":  str(int(site_id)),
            }
            if xl_dev_ip:
                form_data["device_ip"]       = xl_dev_ip
                form_data["device_username"] = xl_dev_user
                form_data["device_password"] = xl_dev_pass
            if xl_dev_vendor:
                form_data["device_vendor"] = xl_dev_vendor
            if xl_dev_model:
                form_data["device_model"] = xl_dev_model
            if xl_dev_info.get("mac"):
                form_data["device_mac"] = xl_dev_info["mac"]
            if xl_dev_info.get("serial"):
                form_data["device_serial"] = xl_dev_info["serial"]
            if company_id_override.strip():
                form_data["company_id"] = company_id_override.strip()

            headers = _backend_headers(auth_token)

            rows = stats["rows"]
            with st.spinner(f"Uploading {rows} user(s) — backend will extract biometrics from device..."):
                try:
                    resp = requests.post(f"{backend_api}/devices/upload-import",
                                         data=form_data, files=files, headers=headers,
                                         timeout=(30, 300), verify=False)
                    if resp.status_code == 201:
                        result = resp.json()
                        st.success("Migration complete!")

                        r1, r2, r3, r4, r5, r6 = st.columns(6)
                        r1.metric("Users Imported",    result.get("imported_user_count", 0))
                        r2.metric("Created",           result.get("created_tenants", 0))
                        r3.metric("Updated",           result.get("updated_tenants", 0))
                        r4.metric("Fingers Stored",    result.get("imported_fingerprint_count", 0))
                        r5.metric("Faces Stored",      result.get("imported_face_count", 0))
                        r6.metric("Cards/PINs",
                                  result.get("imported_card_count", 0)
                                  + result.get("imported_pin_count", 0))

                        for w in result.get("warnings", []):
                            st.warning(w)

                        with st.expander("Per-user results"):
                            st.dataframe(result.get("users", []), use_container_width=True)

                    else:
                        st.error(f"Upload failed: HTTP {resp.status_code}")
                        try:
                            st.code(str(resp.json().get("detail", resp.text[:500])))
                        except Exception:
                            st.code(resp.text[:500])

                except requests.exceptions.ConnectionError:
                    st.error(f"Cannot reach backend at `{backend_api}`.")
                except requests.exceptions.Timeout:
                    st.error(
                        "Timed out waiting for backend. "
                        "Biometric extraction from large devices can take several minutes. "
                        "Check the database before retrying to avoid duplicates."
                    )
                except Exception as exc:
                    st.error(f"Unexpected error: {exc}")

# ---------------------------------------------------------------------------
# Footer
# ---------------------------------------------------------------------------
st.divider()
st.caption(
    "iGatera Migration Extractor — "
    "run this tool on the same local network as your Matrix device."
)
