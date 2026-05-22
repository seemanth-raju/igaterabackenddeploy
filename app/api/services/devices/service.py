import hashlib
import uuid
from datetime import datetime, timezone
from uuid import UUID

from fastapi import HTTPException, status
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.services.companies.service import ensure_company_device_quota, ensure_company_user_quota
from app.api.services.groups.service import validate_group_selection
from app.api.services.devices.schema import DeviceCreate, DeviceImportRequest, DeviceUpdate
from app.core.config import settings
from app.core.security import encrypt_password
from app.services.matrix import MatrixDeviceClient, validate_device_target
from app.utils import get_fingerprint_storage_path
from database.models import (
    AppUser,
    Company,
    Credential,
    Device,
    DeviceUserMapping,
    Site,
    Tenant,
    TenantDeviceAccess,
    TenantSiteAccess,
    UserRole,
)

_DEVICE_MANAGER_ROLES = {
    UserRole.super_admin.value,
    UserRole.company_admin.value,
}


def _hash_push_token(token: str) -> str:
    return hashlib.sha256(token.encode()).hexdigest()


def _normalize_mac(mac_address: str | None) -> str | None:
    if mac_address is None:
        return None
    cleaned = "".join(ch for ch in mac_address if ch.isalnum()).upper()
    if not cleaned:
        return None
    if len(cleaned) != 12 or any(ch not in "0123456789ABCDEF" for ch in cleaned):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid MAC address format")
    return ":".join(cleaned[index:index + 2] for index in range(0, 12, 2))


def _default_serial_number(device_serial_number: str | None, mac_address: str | None) -> str:
    if device_serial_number:
        return device_serial_number
    if mac_address:
        return mac_address.replace(":", "")
    return str(uuid.uuid4())


def _ensure_device_manager(current_user: AppUser) -> None:
    if current_user.role not in _DEVICE_MANAGER_ROLES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")


def resolve_company_id(requested: UUID | None, current_user: AppUser) -> UUID:
    if current_user.role == UserRole.super_admin.value and requested is not None:
        return requested
    return current_user.company_id


def resolve_upload_import_company_id(
    requested_company_id: UUID | None,
    current_user: AppUser | None,
    db: Session,
) -> UUID:
    """Resolve target company for Streamlit upload imports.

    Authenticated uploads keep the existing company-scoped behavior.
    Anonymous uploads are opt-in and meant only for the local extractor flow.
    """
    if current_user is not None:
        _ensure_device_manager(current_user)
        return resolve_company_id(requested_company_id, current_user)

    if not settings.allow_anonymous_migration_uploads:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Migration upload requires authentication",
        )

    if requested_company_id is not None:
        return requested_company_id

    if settings.anonymous_migration_default_company_id is not None:
        return settings.anonymous_migration_default_company_id

    active_companies = [company for company in db.query(Company).all() if company.is_active]
    if len(active_companies) == 1:
        return active_companies[0].company_id
    if not active_companies:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No active company is available for anonymous migration uploads",
        )

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=(
            "company_id is required for anonymous migration uploads when multiple active companies exist"
        ),
    )


def _check_company_active(company_id: UUID, db: Session) -> None:
    company = db.query(Company).filter(Company.company_id == company_id).first()
    if company and not company.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Company is inactive")


def _validate_device_target_or_400(ip_address: str | None, api_port: int | None) -> None:
    if not ip_address:
        return
    try:
        validate_device_target(ip_address, api_port or 80)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


def _resolve_site_id(site_id: int | None, company_id: UUID, db: Session) -> int | None:
    if site_id is None:
        return None
    site = db.query(Site).filter(Site.site_id == site_id).first()
    if not site:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Site not found")
    if site.company_id != company_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Site belongs to another company")
    if not site.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Site is inactive")
    return site.site_id


def _build_matrix_client(payload: DeviceImportRequest) -> MatrixDeviceClient:
    return MatrixDeviceClient(
        device_ip=payload.ip_address,
        username=payload.api_username or "admin",
        password=payload.api_password,
        use_https=bool(payload.use_https),
        api_port=payload.api_port or 80,
    )


def _find_existing_device(company_id: UUID, mac_address: str | None, device_serial_number: str, db: Session) -> Device | None:
    device_by_mac = None
    if mac_address:
        device_by_mac = db.query(Device).filter(Device.mac_address == mac_address).first()
        if device_by_mac and device_by_mac.company_id != company_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="A device with this MAC address already belongs to another company",
            )

    device_by_serial = db.query(Device).filter(Device.device_serial_number == device_serial_number).first()
    if device_by_serial and device_by_serial.company_id != company_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="A device with this serial number already belongs to another company",
        )

    if device_by_mac and device_by_serial and device_by_mac.device_id != device_by_serial.device_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="MAC address and device serial number point to different existing devices",
        )

    return device_by_mac or device_by_serial


def _parse_valid_till(profile: dict) -> datetime | None:
    if (profile.get("validity_enable") or "").strip() != "1":
        return None

    day = (profile.get("validity_date_dd") or "").strip()
    month = (profile.get("validity_date_mm") or "").strip()
    year = (profile.get("validity_date_yyyy") or "").strip()
    if not day or not month or not year:
        return None

    try:
        return datetime(int(year), int(month), int(day), 23, 59, 59, tzinfo=timezone.utc)
    except ValueError:
        return None


def _sanitize_name(profile: dict) -> str:
    name = (profile.get("name") or "").strip()
    fallback = (
        (profile.get("ref_user_id") or "").strip()
        or (profile.get("user_id") or "").strip()
        or (profile.get("user_index") or "").strip()
        or "Imported User"
    )
    return (name or fallback)[:15]


def _tenant_has_changes(tenant: Tenant, profile: dict) -> bool:
    """Return True if any tracked field differs between the existing tenant and the incoming profile."""
    return (
        tenant.full_name != _sanitize_name(profile)
        or tenant.is_active != ((profile.get("user_active") or "1").strip() != "0")
        or tenant.global_access_till != _parse_valid_till(profile)
    )


def _find_tenant_for_import(company_id: UUID, device_id: int, matrix_user_id: str, external_id: str | None, db: Session) -> Tenant | None:
    existing_mapping = (
        db.query(DeviceUserMapping)
        .filter(
            DeviceUserMapping.device_id == device_id,
            DeviceUserMapping.matrix_user_id == matrix_user_id,
        )
        .first()
    )
    if existing_mapping:
        return db.query(Tenant).filter(Tenant.tenant_id == existing_mapping.tenant_id).first()

    if external_id:
        return (
            db.query(Tenant)
            .filter(
                Tenant.company_id == company_id,
                Tenant.external_id == external_id,
            )
            .first()
        )

    return None


def _upsert_tenant_for_import(
    company_id: UUID,
    group_id: int,
    device_id: int,
    profile: dict,
    db: Session,
    on_duplicate: str = "replace",
) -> tuple[Tenant, bool, bool]:
    """Upsert a tenant for import. Returns (tenant, created, skipped).

    on_duplicate controls what happens when the tenant already exists:
      "replace"            — always overwrite profile fields (default)
      "skip"               — leave existing tenant untouched
      "replace_if_changed" — only overwrite if name/active/validity changed
    skipped=True means the tenant existed and its profile was NOT updated.
    """
    matrix_user_id = (
        (profile.get("user_id") or "").strip()
        or (profile.get("ref_user_id") or "").strip()
    )
    external_id = (
        (profile.get("ref_user_id") or "").strip()
        or (profile.get("user_id") or "").strip()
        or None
    )
    tenant = _find_tenant_for_import(company_id, device_id, matrix_user_id, external_id, db)
    valid_till = _parse_valid_till(profile)
    is_active = (profile.get("user_active") or "1").strip() != "0"
    full_name = _sanitize_name(profile)

    if tenant is None:
        ensure_company_user_quota(company_id, db)
        tenant = Tenant(
            company_id=company_id,
            group_id=group_id,
            external_id=external_id,
            full_name=full_name,
            tenant_type="employee",
            is_active=is_active,
            global_access_till=valid_till,
            is_access_enabled=is_active,
        )
        db.add(tenant)
        db.flush()
        return tenant, True, False

    if tenant.company_id != company_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tenant {tenant.tenant_id} belongs to another company",
        )

    if on_duplicate == "skip":
        return tenant, False, True

    if on_duplicate == "replace_if_changed" and not _tenant_has_changes(tenant, profile):
        return tenant, False, True

    if external_id and tenant.external_id != external_id:
        tenant.external_id = external_id
    tenant.full_name = full_name
    tenant.is_active = is_active
    tenant.is_access_enabled = is_active
    tenant.global_access_till = valid_till
    tenant.group_id = group_id
    return tenant, False, False


def _upsert_mapping_for_import(
    tenant: Tenant,
    device: Device,
    profile: dict,
    db: Session,
) -> tuple[DeviceUserMapping, bool]:
    matrix_user_id = (
        (profile.get("user_id") or "").strip()
        or (profile.get("ref_user_id") or "").strip()
    )
    if not matrix_user_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Imported profile is missing user-id")

    mapping = (
        db.query(DeviceUserMapping)
        .filter(
            DeviceUserMapping.tenant_id == tenant.tenant_id,
            DeviceUserMapping.device_id == device.device_id,
        )
        .first()
    )
    created = False
    now = datetime.now(timezone.utc)
    device_response = {
        "imported_from_device": True,
        "user_index": (profile.get("user_index") or "").strip() or None,
        "ref_user_id": (profile.get("ref_user_id") or "").strip() or None,
    }

    if mapping is None:
        mapping = DeviceUserMapping(
            tenant_id=tenant.tenant_id,
            device_id=device.device_id,
            matrix_user_id=matrix_user_id,
            valid_till=_parse_valid_till(profile),
            is_synced=True,
            last_sync_at=now,
            last_sync_attempt_at=now,
            sync_attempt_count=1,
            sync_error=None,
            device_response=device_response,
        )
        db.add(mapping)
        db.flush()
        created = True
        return mapping, created

    mapping.matrix_user_id = matrix_user_id
    mapping.valid_till = _parse_valid_till(profile)
    mapping.is_synced = True
    mapping.last_sync_at = now
    mapping.last_sync_attempt_at = now
    mapping.sync_attempt_count = (mapping.sync_attempt_count or 0) + 1
    mapping.sync_error = None
    mapping.device_response = {**(mapping.device_response or {}), **device_response}
    return mapping, created


def _upsert_site_access_for_import(
    tenant_id: int,
    site_id: int,
    valid_till: datetime | None,
    db: Session,
) -> tuple[TenantSiteAccess, bool]:
    site_access = (
        db.query(TenantSiteAccess)
        .filter(
            TenantSiteAccess.tenant_id == tenant_id,
            TenantSiteAccess.site_id == site_id,
        )
        .first()
    )
    created = False

    if site_access is None:
        site_access = TenantSiteAccess(
            tenant_id=tenant_id,
            site_id=site_id,
            valid_till=valid_till,
        )
        db.add(site_access)
        db.flush()
        created = True
        return site_access, created

    site_access.valid_till = valid_till
    return site_access, created


def _upsert_device_access_for_import(
    tenant_id: int,
    device_id: int,
    site_access_id: int | None,
    valid_till: datetime | None,
    db: Session,
) -> tuple[TenantDeviceAccess, bool]:
    device_access = (
        db.query(TenantDeviceAccess)
        .filter(
            TenantDeviceAccess.tenant_id == tenant_id,
            TenantDeviceAccess.device_id == device_id,
        )
        .first()
    )
    created = False

    if device_access is None:
        device_access = TenantDeviceAccess(
            tenant_id=tenant_id,
            device_id=device_id,
            site_access_id=site_access_id,
            valid_till=valid_till,
        )
        db.add(device_access)
        db.flush()
        created = True
        return device_access, created

    device_access.site_access_id = site_access_id
    device_access.valid_till = valid_till
    return device_access, created


def _upsert_imported_fingerprint(
    tenant_id: int,
    finger_index: int,
    template_bytes: bytes,
    db: Session,
) -> tuple[Credential, bool]:
    storage_path = get_fingerprint_storage_path()
    file_path = storage_path / f"tenant_{tenant_id}_finger_{finger_index}.dat"
    file_path.write_bytes(template_bytes)
    file_hash = hashlib.sha256(template_bytes).hexdigest()

    credential = (
        db.query(Credential)
        .filter(
            Credential.tenant_id == tenant_id,
            Credential.type == "finger",
            Credential.slot_index == finger_index,
        )
        .first()
    )
    created = False

    if credential is None:
        credential = Credential(
            tenant_id=tenant_id,
            type="finger",
            slot_index=finger_index,
            file_path=str(file_path),
            file_hash=file_hash,
            algorithm_version="matrix_v1",
        )
        db.add(credential)
        db.flush()
        created = True
        return credential, created

    credential.file_path = str(file_path)
    credential.file_hash = file_hash
    credential.algorithm_version = "matrix_v1"
    return credential, created


def _upsert_imported_face(
    tenant_id: int,
    face_no: int,
    template_bytes: bytes,
    db: Session,
) -> Credential:
    storage_path = get_fingerprint_storage_path()
    file_path = storage_path / f"tenant_{tenant_id}_face_{face_no}.dat"
    file_path.write_bytes(template_bytes)
    file_hash = hashlib.sha256(template_bytes).hexdigest()

    credential = (
        db.query(Credential)
        .filter(
            Credential.tenant_id == tenant_id,
            Credential.type == "face",
            Credential.slot_index == face_no,
        )
        .first()
    )
    if credential is None:
        credential = Credential(
            tenant_id=tenant_id,
            type="face",
            slot_index=face_no,
            file_path=str(file_path),
            file_hash=file_hash,
            algorithm_version="matrix_v1",
        )
        db.add(credential)
    else:
        credential.file_path = str(file_path)
        credential.file_hash = file_hash
        credential.algorithm_version = "matrix_v1"
    db.flush()
    return credential


def _upsert_imported_card(tenant_id: int, card_value: str, slot: int, db: Session) -> None:
    credential = (
        db.query(Credential)
        .filter(
            Credential.tenant_id == tenant_id,
            Credential.type == "card",
            Credential.slot_index == slot,
        )
        .first()
    )
    if credential is None:
        db.add(Credential(tenant_id=tenant_id, type="card", slot_index=slot, raw_value=card_value))
    else:
        credential.raw_value = card_value
    db.flush()


def _upsert_imported_pin(tenant_id: int, pin_value: str, db: Session) -> None:
    credential = (
        db.query(Credential)
        .filter(Credential.tenant_id == tenant_id, Credential.type == "pin")
        .first()
    )
    if credential is None:
        db.add(Credential(tenant_id=tenant_id, type="pin", slot_index=1, raw_value=pin_value))
    else:
        credential.raw_value = pin_value
    db.flush()


def _import_face_for_import(
    client: MatrixDeviceClient,
    tenant_id: int,
    matrix_user_id: str,
    mapping: DeviceUserMapping,
    db: Session,
) -> int:
    template_bytes = client.get_face_template(matrix_user_id, face_no=1)
    if not template_bytes:
        return 0
    _upsert_imported_face(tenant_id, face_no=1, template_bytes=template_bytes, db=db)
    mapping.device_response = {
        **(mapping.device_response or {}),
        "face_imported": True,
    }
    return 1


def _import_fingers_from_device(
    client: MatrixDeviceClient,
    tenant_id: int,
    matrix_user_id: str,
    mapping: DeviceUserMapping,
    db: Session,
) -> int:
    imported = 0
    for finger_index, template_bytes in client.list_fingerprint_templates(matrix_user_id).items():
        if not template_bytes:
            continue
        _upsert_imported_fingerprint(tenant_id, finger_index, template_bytes, db)
        imported += 1
    if imported:
        mapping.device_response = {
            **(mapping.device_response or {}),
            "imported_finger_count": imported,
            "fingerprint_imported": True,
        }
    return imported


def _import_fingerprints_for_import(
    client: MatrixDeviceClient,
    tenant_id: int,
    matrix_user_id: str,
    mapping: DeviceUserMapping,
    db: Session,
) -> int:
    if not matrix_user_id:
        return 0

    imported = 0
    for finger_index, template_bytes in client.list_fingerprint_templates(matrix_user_id).items():
        if not template_bytes:
            continue
        _upsert_imported_fingerprint(tenant_id, finger_index, template_bytes, db)
        imported += 1

    mapping.device_response = {
        **(mapping.device_response or {}),
        "imported_finger_count": imported,
        "fingerprint_imported": imported > 0,
    }
    return imported


def create_device(payload: DeviceCreate, company_id: UUID, db: Session) -> Device:
    _check_company_active(company_id, db)
    _validate_device_target_or_400(payload.ip_address, payload.api_port)
    ensure_company_device_quota(company_id, db)
    normalized_mac = _normalize_mac(payload.mac_address)
    device_serial_number = _default_serial_number(payload.device_serial_number, normalized_mac)
    encrypted_password = encrypt_password(payload.api_password) if payload.api_password else None
    push_token_hash = _hash_push_token(payload.push_token) if payload.push_token else None

    device = Device(
        company_id=company_id,
        site_id=_resolve_site_id(payload.site_id, company_id, db),
        device_serial_number=device_serial_number,
        vendor=payload.vendor,
        model_name=payload.model_name,
        ip_address=payload.ip_address,
        mac_address=normalized_mac,
        api_username=payload.api_username,
        api_password_encrypted=encrypted_password,
        api_port=payload.api_port,
        use_https=payload.use_https,
        is_active=payload.is_active,
        communication_mode=payload.communication_mode,
        push_token_hash=push_token_hash,
        status=payload.status,
        config=payload.config,
        credential_types=payload.credential_types,
    )
    db.add(device)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Device with same MAC or serial number already exists",
        ) from exc
    db.refresh(device)
    return device


def list_devices(
    db: Session,
    company_id: UUID | None = None,
    site_id: int | None = None,
    skip: int = 0,
    limit: int = 50,
    search: str | None = None,
) -> list[Device]:
    query = db.query(Device)

    if company_id is not None:
        query = query.filter(Device.company_id == company_id)

    if site_id is not None:
        query = query.filter(Device.site_id == site_id)

    if search:
        like_value = f"%{search}%"
        query = query.filter(
            Device.device_serial_number.ilike(like_value)
            | Device.vendor.ilike(like_value)
            | Device.model_name.ilike(like_value)
            | Device.ip_address.ilike(like_value)
        )

    return query.order_by(Device.created_at.desc()).offset(skip).limit(limit).all()


def get_device(device_id: int, db: Session) -> Device:
    device = db.query(Device).filter(Device.device_id == device_id).first()
    if not device:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Device not found")
    return device


def update_device(device_id: int, payload: DeviceUpdate, db: Session) -> Device:
    device = get_device(device_id, db)
    updated = payload.model_fields_set

    if "site_id" in updated:
        device.site_id = _resolve_site_id(payload.site_id, device.company_id, db)
    if "device_serial_number" in updated:
        device.device_serial_number = _default_serial_number(payload.device_serial_number, device.mac_address)
    if "vendor" in updated:
        device.vendor = payload.vendor
    if "model_name" in updated:
        device.model_name = payload.model_name
    if "ip_address" in updated:
        _validate_device_target_or_400(payload.ip_address, payload.api_port if "api_port" in updated else device.api_port)
        device.ip_address = payload.ip_address
    if "mac_address" in updated:
        previous_mac = device.mac_address
        previous_serial = device.device_serial_number
        normalized_mac = _normalize_mac(payload.mac_address)
        device.mac_address = normalized_mac
        if (
            "device_serial_number" not in updated
            and previous_mac
            and previous_serial == previous_mac.replace(":", "")
        ):
            device.device_serial_number = _default_serial_number(None, normalized_mac)
    if "api_username" in updated:
        device.api_username = payload.api_username
    if "api_password" in updated:
        device.api_password_encrypted = encrypt_password(payload.api_password) if payload.api_password else None
    if "api_port" in updated:
        _validate_device_target_or_400(device.ip_address, payload.api_port)
        device.api_port = payload.api_port
    if "use_https" in updated:
        device.use_https = payload.use_https
    if "is_active" in updated:
        device.is_active = payload.is_active
    if "communication_mode" in updated:
        device.communication_mode = payload.communication_mode
    if "push_token" in updated:
        device.push_token_hash = _hash_push_token(payload.push_token) if payload.push_token else None
    if "status" in updated:
        device.status = payload.status
    if "config" in updated:
        device.config = payload.config
    if "credential_types" in updated:
        device.credential_types = payload.credential_types

    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Device with same MAC or serial number already exists",
        ) from exc
    db.refresh(device)
    return device


def import_enrollment_device(payload: DeviceImportRequest, current_user: AppUser, db: Session) -> dict:
    _ensure_device_manager(current_user)

    company_id = resolve_company_id(payload.company_id, current_user)
    _check_company_active(company_id, db)
    _validate_device_target_or_400(payload.ip_address, payload.api_port)
    group = validate_group_selection(company_id, payload.group_id, db)

    client = _build_matrix_client(payload)
    if not client.ping():
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Could not reach device at {payload.ip_address}:{payload.api_port or 80}",
        )

    warnings: list[str] = []
    reported_user_count = client.get_user_count()
    profiles = client.list_user_profiles(reported_user_count if reported_user_count >= 0 else None)
    if reported_user_count < 0:
        if not profiles:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="Device responded, but user records could not be read. Check credentials and API access.",
            )
        warnings.append(
            "Device user count endpoint was unavailable; users were imported by scanning device records."
        )
        reported_user_count = len(profiles)

    if reported_user_count != len(profiles):
        warnings.append(
            f"Device reported {reported_user_count} user(s), but {len(profiles)} profile(s) were imported."
        )

    normalized_mac = _normalize_mac(payload.mac_address)
    device_serial_number = _default_serial_number(payload.device_serial_number, normalized_mac)
    existing_device = _find_existing_device(company_id, normalized_mac, device_serial_number, db)
    encrypted_password = encrypt_password(payload.api_password)
    site_id = _resolve_site_id(payload.site_id, company_id, db)
    now = datetime.now(timezone.utc)
    vendor = payload.vendor or "Matrix"
    model_name = payload.model_name or "COSEC"
    api_username = payload.api_username or "admin"
    api_port = payload.api_port or 80
    use_https = bool(payload.use_https)
    communication_mode = payload.communication_mode or "direct"

    if existing_device is None:
        ensure_company_device_quota(company_id, db)
        device = Device(
            company_id=company_id,
            site_id=site_id,
            device_serial_number=device_serial_number,
            vendor=vendor,
            model_name=model_name,
            ip_address=payload.ip_address,
            mac_address=normalized_mac,
            api_username=api_username,
            api_password_encrypted=encrypted_password,
            api_port=api_port,
            use_https=use_https,
            is_active=True,
            communication_mode=communication_mode,
            status="online",
            last_heartbeat=now,
            config={"last_import": {"source": "enrollment_device", "at": now.isoformat()}},
        )
        db.add(device)
        db.flush()
        device_created = True
    else:
        device = existing_device
        device.site_id = site_id
        if payload.vendor is not None:
            device.vendor = vendor
        if payload.model_name is not None:
            device.model_name = model_name
        device.ip_address = payload.ip_address
        device.mac_address = normalized_mac
        if payload.api_username is not None:
            device.api_username = api_username
        device.api_password_encrypted = encrypted_password
        if payload.api_port is not None:
            device.api_port = api_port
        if payload.use_https is not None:
            device.use_https = use_https
        device.is_active = True
        if payload.communication_mode is not None:
            device.communication_mode = communication_mode
        device.status = "online"
        device.last_heartbeat = now
        device.config = {
            **(device.config or {}),
            "last_import": {"source": "enrollment_device", "at": now.isoformat()},
        }
        device_created = False

    users: list[dict] = []
    created_tenants = 0
    updated_tenants = 0
    skipped_users = 0
    created_mappings = 0
    updated_mappings = 0
    created_device_accesses = 0
    created_site_accesses = 0
    imported_fingerprint_count = 0
    users_with_fingerprints = 0

    for profile in profiles:
        tenant, tenant_created, tenant_skipped = _upsert_tenant_for_import(
            company_id=company_id,
            group_id=group.group_id,
            device_id=device.device_id,
            profile=profile,
            db=db,
        )
        mapping, mapping_created = _upsert_mapping_for_import(tenant, device, profile, db)
        site_access, site_access_created = _upsert_site_access_for_import(
            tenant_id=tenant.tenant_id,
            site_id=device.site_id,
            valid_till=_parse_valid_till(profile),
            db=db,
        )
        _, device_access_created = _upsert_device_access_for_import(
            tenant_id=tenant.tenant_id,
            device_id=device.device_id,
            site_access_id=site_access.site_access_id,
            valid_till=_parse_valid_till(profile),
            db=db,
        )
        imported_fingers = _import_fingerprints_for_import(
            client=client,
            tenant_id=tenant.tenant_id,
            matrix_user_id=mapping.matrix_user_id,
            mapping=mapping,
            db=db,
        )

        created_tenants += 1 if tenant_created else 0
        updated_tenants += 1 if (not tenant_created and not tenant_skipped) else 0
        skipped_users += 1 if tenant_skipped else 0
        created_mappings += 1 if mapping_created else 0
        updated_mappings += 0 if mapping_created else 1
        created_site_accesses += 1 if site_access_created else 0
        created_device_accesses += 1 if device_access_created else 0
        imported_fingerprint_count += imported_fingers
        users_with_fingerprints += 1 if imported_fingers > 0 else 0

        users.append({
            "tenant_id": tenant.tenant_id,
            "matrix_user_id": mapping.matrix_user_id,
            "external_id": tenant.external_id,
            "full_name": tenant.full_name,
            "is_active": tenant.is_active,
            "valid_till": mapping.valid_till,
            "finger_count": imported_fingers,
            "face_count": 0,
            "card_count": 0,
            "has_pin": False,
            "tenant_created": tenant_created,
            "mapping_created": mapping_created,
            "tenant_skipped": tenant_skipped,
        })

    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not persist imported device data",
        ) from exc

    db.refresh(device)
    db.refresh(group)

    return {
        "device": device,
        "device_created": device_created,
        "group": group,
        "reported_user_count": reported_user_count,
        "imported_user_count": len(users),
        "created_tenants": created_tenants,
        "updated_tenants": updated_tenants,
        "skipped_users": skipped_users,
        "created_mappings": created_mappings,
        "updated_mappings": updated_mappings,
        "created_device_accesses": created_device_accesses,
        "created_site_accesses": created_site_accesses,
        "imported_fingerprint_count": imported_fingerprint_count,
        "users_with_fingerprints": users_with_fingerprints,
        "imported_face_count": 0,
        "users_with_faces": 0,
        "imported_card_count": 0,
        "imported_pin_count": 0,
        "warnings": warnings,
        "users": users,
    }


def delete_device(device_id: int, db: Session) -> None:
    device = get_device(device_id, db)
    db.delete(device)
    db.commit()


# ---------------------------------------------------------------------------
# Upload-based import (used by Streamlit extractor running on the local LAN)
# ---------------------------------------------------------------------------

def _parse_excel_profiles(excel_bytes: bytes, filename: str = "") -> list[dict]:
    """Parse an Excel file into profile dicts compatible with _upsert_tenant_for_import.

    Supports two formats:
    - Generic:  user_id, full_name, ref_user_id, is_active, valid_till, user_index
    - Matrix export (User_Configuration.xls):
        User ID, User Name, Active, Valid Upto (dd-mm-yyyy),
        Card1, Card2, PIN, Face Recognition
    All header matching is case-insensitive, spaces/dashes normalised to underscores.
    Supports both .xls (xlrd) and .xlsx (openpyxl).
    """
    import io as _io

    is_xls = filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx")

    try:
        if is_xls:
            import xlrd
            wb_xls = xlrd.open_workbook(file_contents=excel_bytes)
            ws_xls = wb_xls.sheet_by_index(0)
            all_rows: list = [ws_xls.row_values(i) for i in range(ws_xls.nrows)]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(_io.BytesIO(excel_bytes), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot parse Excel file: {exc}",
        ) from exc

    headers: list[str] = []
    profiles: list[dict] = []

    for row_idx, row in enumerate(all_rows):
        if row_idx == 0:
            headers = [
                str(c).strip().lower().replace(" ", "_").replace("-", "_") if c else ""
                for c in row
            ]
            continue
        if not any(row):
            continue

        row_dict = {
            headers[i]: (str(v).strip() if v is not None else "")
            for i, v in enumerate(row)
            if i < len(headers) and headers[i]
        }

        # user_id — Matrix column is "User ID" → normalised to "user_id"
        user_id = row_dict.get("user_id", "").strip()
        if not user_id:
            continue

        # full_name — generic "full_name"/"name" OR Matrix "user_name"
        full_name = (
            row_dict.get("full_name")
            or row_dict.get("name")
            or row_dict.get("user_name")
            or user_id
        ).strip()

        # active — generic "is_active" OR Matrix "active"
        is_active_raw = (row_dict.get("is_active") or row_dict.get("active") or "1").lower()
        user_active = "0" if is_active_raw in ("0", "false", "no", "inactive") else "1"

        # valid_till — generic "valid_till" (YYYY-MM-DD) OR Matrix "valid_upto" (dd-mm-yyyy,
        # sometimes prefixed with a backtick as an Excel text-escape artefact)
        validity_extras: dict = {}
        vt_raw = (row_dict.get("valid_till") or row_dict.get("valid_upto") or "").strip().lstrip("`")
        if vt_raw and vt_raw not in ("None", "nan"):
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    dt = datetime.strptime(vt_raw[:10], fmt)
                    validity_extras = {
                        "validity_enable": "1",
                        "validity_date_dd": str(dt.day),
                        "validity_date_mm": str(dt.month),
                        "validity_date_yyyy": str(dt.year),
                    }
                    break
                except ValueError:
                    continue

        # credential fields from Matrix export
        card1 = row_dict.get("card1", "").strip()
        card2 = row_dict.get("card2", "").strip()
        pin   = row_dict.get("pin", "").strip()
        # "Face Recognition" → "face_recognition" after normalisation
        has_face = "1" if row_dict.get("face_recognition", "0").strip() == "1" else "0"

        profiles.append({
            "user_id": user_id,
            "ref_user_id": row_dict.get("ref_user_id", user_id) or user_id,
            "user_index": row_dict.get("user_index", ""),
            "name": full_name,
            "user_active": user_active,
            "validity_enable": validity_extras.get("validity_enable", "0"),
            "validity_date_dd": validity_extras.get("validity_date_dd", ""),
            "validity_date_mm": validity_extras.get("validity_date_mm", ""),
            "validity_date_yyyy": validity_extras.get("validity_date_yyyy", ""),
            "card1": card1,
            "card2": card2,
            "pin": pin,
            "has_face": has_face,
        })

    return profiles


def _parse_fingerprint_filename(filename: str) -> tuple[str, int] | tuple[None, None]:
    """Parse '{user_id}_finger_{index}.dat' → (user_id, finger_index).

    Examples: '42_finger_1.dat', 'USR001_finger_2.dat'
    """
    import re
    m = re.match(r"^(.+)_finger_(\d+)\.dat$", filename, re.IGNORECASE)
    if m:
        return m.group(1), int(m.group(2))
    return None, None


def import_from_upload(
    group_id: int,
    site_id: int,
    company_id: UUID,
    excel_bytes: bytes,
    excel_filename: str,
    fingerprint_files: list[tuple[str, bytes]],
    device_ip: str | None,
    device_mac: str | None,
    device_serial: str | None,
    device_vendor: str | None,
    device_model: str | None,
    device_username: str | None,
    device_password: str | None,
    db: Session,
    on_duplicate: str = "replace_if_changed",
) -> dict:
    """Import tenants from an uploaded Excel + optional direct device biometric extraction.

    Supports the Matrix User_Configuration.xls export format (28 columns) as well as the
    generic format (user_id, full_name, is_active, valid_till).

    Credentials are resolved automatically from the Excel and device:
      - Card1 / Card2 / PIN  → read directly from Excel columns (no device call)
      - Face template        → fetched from device when Face Recognition = 1 in Excel
      - Finger templates     → fetched from device when credentials provided;
                               falls back to uploaded .dat files if no device connection
    """
    from app.api.services.groups.service import validate_group_selection

    _check_company_active(company_id, db)
    group = validate_group_selection(company_id, group_id, db)

    profiles = _parse_excel_profiles(excel_bytes, filename=excel_filename)
    if not profiles:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file is empty or contains no valid user rows (check that 'user_id' column is present).",
        )

    # Build fingerprint lookup from uploaded .dat files (fallback if device unreachable)
    fingerprint_map: dict[str, dict[int, bytes]] = {}
    for filename, content in fingerprint_files:
        uid, finger_index = _parse_fingerprint_filename(filename)
        if uid is not None and finger_index is not None and content:
            fingerprint_map.setdefault(uid, {})[finger_index] = content

    # Try to build a direct device client for biometric extraction
    warnings: list[str] = []
    client: MatrixDeviceClient | None = None
    if device_ip and device_username and device_password:
        client = MatrixDeviceClient(
            device_ip=device_ip,
            username=device_username,
            password=device_password,
            use_https=False,
        )
        if not client.ping():
            client = None
            warnings.append(
                f"Could not reach device at {device_ip} — biometric extraction from device skipped."
                + (" Finger templates from uploaded .dat files will be used instead." if fingerprint_map else "")
            )

    # Resolve site (required — all tenants get site access)
    site_id_resolved = _resolve_site_id(site_id, company_id, db)

    # Resolve or create device record (for DeviceUserMapping)
    normalized_mac = _normalize_mac(device_mac) if device_mac else None
    device_serial_number = _default_serial_number(device_serial, normalized_mac)
    existing_device = _find_existing_device(company_id, normalized_mac, device_serial_number, db)
    now = datetime.now(timezone.utc)

    if existing_device is None:
        ensure_company_device_quota(company_id, db)
        device = Device(
            company_id=company_id,
            site_id=site_id_resolved,
            device_serial_number=device_serial_number,
            vendor=device_vendor or "Matrix",
            model_name=device_model or "COSEC",
            ip_address=device_ip,
            mac_address=normalized_mac,
            is_active=True,
            communication_mode="direct",
            status="unknown",
            config={"last_import": {"source": "streamlit_upload", "at": now.isoformat()}},
        )
        db.add(device)
        db.flush()
        device_created = True
    else:
        device = existing_device
        device.site_id = site_id_resolved
        if device_ip:
            device.ip_address = device_ip
        if normalized_mac:
            device.mac_address = normalized_mac
        if device_vendor:
            device.vendor = device_vendor
        if device_model:
            device.model_name = device_model
        device.config = {
            **(device.config or {}),
            "last_import": {"source": "streamlit_upload", "at": now.isoformat()},
        }
        device_created = False

    # Process each user row
    users: list[dict] = []
    created_tenants = updated_tenants = skipped_users = 0
    created_mappings = updated_mappings = 0
    created_site_accesses = created_device_accesses = 0
    imported_fingerprint_count = users_with_fingerprints = 0
    imported_face_count = users_with_faces = 0
    imported_card_count = imported_pin_count = 0

    for profile in profiles:
        tenant, tenant_created, tenant_skipped = _upsert_tenant_for_import(
            company_id=company_id,
            group_id=group.group_id,
            device_id=device.device_id,
            profile=profile,
            db=db,
            on_duplicate=on_duplicate,
        )

        # "skip" mode: leave existing users completely untouched
        if tenant_skipped and on_duplicate == "skip":
            skipped_users += 1
            matrix_user_id_raw = (profile.get("user_id") or profile.get("ref_user_id") or "").strip()
            users.append({
                "tenant_id": tenant.tenant_id,
                "matrix_user_id": matrix_user_id_raw,
                "external_id": tenant.external_id,
                "full_name": tenant.full_name,
                "is_active": tenant.is_active,
                "valid_till": None,
                "finger_count": 0,
                "face_count": 0,
                "card_count": 0,
                "has_pin": False,
                "tenant_created": False,
                "mapping_created": False,
                "tenant_skipped": True,
            })
            continue

        mapping, mapping_created = _upsert_mapping_for_import(tenant, device, profile, db)
        site_access, site_access_created = _upsert_site_access_for_import(
            tenant_id=tenant.tenant_id,
            site_id=device.site_id,
            valid_till=_parse_valid_till(profile),
            db=db,
        )
        _, device_access_created = _upsert_device_access_for_import(
            tenant_id=tenant.tenant_id,
            device_id=device.device_id,
            site_access_id=site_access.site_access_id,
            valid_till=_parse_valid_till(profile),
            db=db,
        )

        matrix_user_id = mapping.matrix_user_id

        # Finger templates — device first, fall back to uploaded .dat files
        imported_fingers = 0
        if client:
            imported_fingers = _import_fingers_from_device(
                client, tenant.tenant_id, matrix_user_id, mapping, db
            )
        if not imported_fingers:
            for finger_index, template_bytes in fingerprint_map.get(matrix_user_id, {}).items():
                _upsert_imported_fingerprint(tenant.tenant_id, finger_index, template_bytes, db)
                imported_fingers += 1
            if imported_fingers:
                mapping.device_response = {
                    **(mapping.device_response or {}),
                    "imported_finger_count": imported_fingers,
                    "fingerprint_imported": True,
                }

        # Face template — from device when Excel says Face Recognition = 1
        imported_faces = 0
        if client and profile.get("has_face") == "1":
            imported_faces = _import_face_for_import(
                client, tenant.tenant_id, matrix_user_id, mapping, db
            )

        # Card — read directly from Excel (no device call needed)
        imported_cards = 0
        if profile.get("card1"):
            _upsert_imported_card(tenant.tenant_id, profile["card1"], slot=1, db=db)
            imported_cards += 1
        if profile.get("card2"):
            _upsert_imported_card(tenant.tenant_id, profile["card2"], slot=2, db=db)
            imported_cards += 1

        # PIN — read directly from Excel (no device call needed)
        imported_pin = 0
        if profile.get("pin"):
            _upsert_imported_pin(tenant.tenant_id, profile["pin"], db=db)
            imported_pin = 1

        created_tenants += 1 if tenant_created else 0
        updated_tenants += 1 if (not tenant_created and not tenant_skipped) else 0
        skipped_users += 1 if tenant_skipped else 0
        created_mappings += 1 if mapping_created else 0
        updated_mappings += 0 if mapping_created else 1
        created_site_accesses += 1 if site_access_created else 0
        created_device_accesses += 1 if device_access_created else 0
        imported_fingerprint_count += imported_fingers
        users_with_fingerprints += 1 if imported_fingers else 0
        imported_face_count += imported_faces
        users_with_faces += 1 if imported_faces else 0
        imported_card_count += imported_cards
        imported_pin_count += imported_pin

        users.append({
            "tenant_id": tenant.tenant_id,
            "matrix_user_id": matrix_user_id,
            "external_id": tenant.external_id,
            "full_name": tenant.full_name,
            "is_active": tenant.is_active,
            "valid_till": mapping.valid_till,
            "finger_count": imported_fingers,
            "face_count": imported_faces,
            "card_count": imported_cards,
            "has_pin": bool(imported_pin),
            "tenant_created": tenant_created,
            "mapping_created": mapping_created,
            "tenant_skipped": tenant_skipped,
        })

    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not persist uploaded import data",
        ) from exc

    db.refresh(device)
    db.refresh(group)

    return {
        "device": device,
        "device_created": device_created,
        "group": group,
        "reported_user_count": len(profiles),
        "imported_user_count": len(users),
        "created_tenants": created_tenants,
        "updated_tenants": updated_tenants,
        "skipped_users": skipped_users,
        "created_mappings": created_mappings,
        "updated_mappings": updated_mappings,
        "created_device_accesses": created_device_accesses,
        "created_site_accesses": created_site_accesses,
        "imported_fingerprint_count": imported_fingerprint_count,
        "users_with_fingerprints": users_with_fingerprints,
        "imported_face_count": imported_face_count,
        "users_with_faces": users_with_faces,
        "imported_card_count": imported_card_count,
        "imported_pin_count": imported_pin_count,
        "warnings": warnings,
        "users": users,
    }
