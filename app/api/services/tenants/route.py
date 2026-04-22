from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, get_db
from app.api.services.groups.schema import TenantGroupRead
from app.api.services.groups.service import get_tenant_group, validate_group_selection
from app.api.services.tenants.enrollment import (
    _get_device_or_404,
    enroll_to_device,
    enroll_to_devices_bulk,
    enroll_to_site,
    extract_fingerprint_from_device,
    register_and_capture_fingerprint,
    unenroll_from_device,
    unenroll_from_devices_bulk,
    update_device_access_validity,
    update_tenant_on_device,
    update_tenant_on_devices_bulk,
)
from app.api.services.tenants.schema import (
    BulkEnrollRequest,
    CaptureRequest,
    DeviceAccessRead,
    DeviceAccessUpdate,
    DeviceEnrollRequest,
    SiteEnrollRequest,
    TenantCreate,
    TenantRead,
    TenantUpdate,
)
from app.api.services.tenants.service import (
    create_tenant,
    delete_tenant_with_related_data,
    get_tenant,
    list_tenants,
    update_tenant,
)
from database.models import AppUser, Credential, DeviceUserMapping, TenantGroup, UserRole

import logging
_log = logging.getLogger(__name__)

router = APIRouter(prefix="/tenants", tags=["tenants"])

_TENANT_MANAGER_ROLES = {UserRole.super_admin.value, UserRole.company_admin.value}


def _to_tenant_read(tenant, db: Session, *, cred_summary: dict | None = None, device_count: int | None = None) -> TenantRead:
    """Convert a Tenant ORM object to TenantRead.

    For single-tenant lookups, fetches credentials and device count inline.
    For list operations, pass pre-fetched cred_summary and device_count to avoid N+1.
    """
    if cred_summary is None:
        creds = db.query(Credential.type).filter(Credential.tenant_id == tenant.tenant_id).all()
        cred_types = [c[0] for c in creds]
        cred_summary = {
            "finger_count": sum(1 for t in cred_types if t == "finger"),
            "has_face": "face" in cred_types,
            "has_card": "card" in cred_types,
        }
    if device_count is None:
        device_count = db.query(DeviceUserMapping).filter(DeviceUserMapping.tenant_id == tenant.tenant_id).count()
    group = get_tenant_group(tenant.tenant_id, db)

    return TenantRead(
        tenant_id=tenant.tenant_id,
        company_id=str(tenant.company_id) if tenant.company_id else None,
        external_id=tenant.external_id,
        full_name=tenant.full_name,
        email=tenant.email,
        phone=tenant.phone,
        tenant_type=tenant.tenant_type,
        is_active=tenant.is_active,
        is_access_enabled=tenant.is_access_enabled,
        global_access_from=tenant.global_access_from,
        global_access_till=tenant.global_access_till,
        access_timezone=tenant.access_timezone,
        created_at=tenant.created_at,
        finger_count=cred_summary["finger_count"],
        has_face=cred_summary["has_face"],
        has_card=cred_summary["has_card"],
        enrolled_device_count=device_count,
        group=group,
    )


def _batch_tenant_reads(tenants: list, db: Session) -> list[TenantRead]:
    """Convert a list of tenants to TenantRead with batch queries (avoids N+1)."""
    if not tenants:
        return []

    tenant_ids = [t.tenant_id for t in tenants]

    # Batch fetch credentials
    cred_rows = (
        db.query(Credential.tenant_id, Credential.type)
        .filter(Credential.tenant_id.in_(tenant_ids))
        .all()
    )
    cred_map: dict[int, list[str]] = {}
    for tid, ctype in cred_rows:
        cred_map.setdefault(tid, []).append(ctype)

    # Batch fetch device mapping counts
    from sqlalchemy import func as sa_func
    count_rows = (
        db.query(DeviceUserMapping.tenant_id, sa_func.count(DeviceUserMapping.mapping_id))
        .filter(DeviceUserMapping.tenant_id.in_(tenant_ids))
        .group_by(DeviceUserMapping.tenant_id)
        .all()
    )
    count_map = {tid: cnt for tid, cnt in count_rows}

    group_map: dict[int, TenantGroupRead] = {}
    group_ids = sorted({t.group_id for t in tenants if t.group_id is not None})
    if group_ids:
        groups = db.query(TenantGroup).filter(TenantGroup.group_id.in_(group_ids)).all()
        group_lookup = {
            group.group_id: TenantGroupRead(
                group_id=group.group_id,
                name=group.name,
                code=group.code,
                short_name=group.short_name,
            )
            for group in groups
        }
        group_map = {
            tenant.tenant_id: group_lookup[tenant.group_id]
            for tenant in tenants
            if tenant.group_id is not None and tenant.group_id in group_lookup
        }

    results = []
    for t in tenants:
        types = cred_map.get(t.tenant_id, [])
        summary = {
            "finger_count": sum(1 for tp in types if tp == "finger"),
            "has_face": "face" in types,
            "has_card": "card" in types,
        }
        results.append(
            TenantRead(
                tenant_id=t.tenant_id,
                company_id=str(t.company_id) if t.company_id else None,
                external_id=t.external_id,
                full_name=t.full_name,
                email=t.email,
                phone=t.phone,
                tenant_type=t.tenant_type,
                is_active=t.is_active,
                is_access_enabled=t.is_access_enabled,
                global_access_from=t.global_access_from,
                global_access_till=t.global_access_till,
                access_timezone=t.access_timezone,
                created_at=t.created_at,
                finger_count=summary["finger_count"],
                has_face=summary["has_face"],
                has_card=summary["has_card"],
                enrolled_device_count=count_map.get(t.tenant_id, 0),
                group=group_map.get(t.tenant_id),
            )
        )
    return results


def _check_tenant_access(tenant, current_user: AppUser) -> None:
    if current_user.role != UserRole.super_admin.value and tenant.company_id != current_user.company_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized to access this tenant")


def _require_tenant_manager(current_user: AppUser) -> None:
    if current_user.role not in _TENANT_MANAGER_ROLES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")


def _check_device_company_access(device, current_user: AppUser) -> None:
    if current_user.role != UserRole.super_admin.value and device.company_id != current_user.company_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized to access this device")


def _resolve_company_id(requested: UUID | None, current_user: AppUser) -> UUID:
    if current_user.role == UserRole.super_admin.value and requested is not None:
        return requested
    return current_user.company_id


# ---------------------------------------------------------------------------
# CRUD  (Page 1 — Add/manage users)
# ---------------------------------------------------------------------------


@router.post("", response_model=TenantRead, status_code=status.HTTP_201_CREATED)
def create_tenant_route(
    payload: TenantCreate,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> TenantRead:
    """Create a new tenant — basic details only, no device interaction."""
    _require_tenant_manager(current_user)
    company_id = _resolve_company_id(payload.company_id, current_user)
    tenant = create_tenant(payload, company_id, db)
    return _to_tenant_read(tenant, db)


@router.get("", response_model=list[TenantRead])
def list_tenants_route(
    company_id: UUID | None = Query(default=None),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    search: str | None = Query(default=None),
    group_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> list[TenantRead]:
    if current_user.role != UserRole.super_admin.value:
        company_id = current_user.company_id
    tenants = list_tenants(
        db,
        company_id=company_id,
        skip=skip,
        limit=limit,
        search=search,
        group_id=group_id,
    )
    return _batch_tenant_reads(tenants, db)


# ---------------------------------------------------------------------------
# Bulk import / template  (must be before /{tenant_id} to avoid path clash)
# ---------------------------------------------------------------------------

TEMPLATE_COLUMNS = [
    "company_id",
    "group_id*",
    "external_id",
    "full_name*",
    "email",
    "phone",
    "tenant_type",
    "is_active",
    "global_access_from",
    "global_access_till",
]


@router.get("/template")
def download_template(
    company_id: UUID | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    """Download an Excel template for bulk tenant import."""
    _require_tenant_manager(current_user)
    resolved_company_id = _resolve_company_id(company_id, current_user)
    if resolved_company_id is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="company_id is required")
    wb = Workbook()
    ws = wb.active
    ws.title = "Tenants"

    # Header row
    for col_idx, header in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    # Example row
    default_group = (
        db.query(TenantGroup)
        .filter(TenantGroup.company_id == resolved_company_id, TenantGroup.is_active == True)
        .order_by(TenantGroup.is_default.desc(), TenantGroup.name.asc())
        .first()
    )
    example = [
        str(resolved_company_id) if current_user.role == UserRole.super_admin.value else "",
        default_group.group_id if default_group else "",
        "EMP001",
        "John Doe",
        "john@example.com",
        "+1-555-0100",
        "employee",
        "true",
        "2026-05-01 09:00",
        "2026-12-31 18:00",
    ]
    for col_idx, val in enumerate(example, 1):
        ws.cell(row=2, column=col_idx, value=val)

    groups_ws = wb.create_sheet("Groups")
    for col_idx, header in enumerate(["group_id", "name", "code", "is_default"], 1):
        cell = groups_ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)
    groups = (
        db.query(TenantGroup)
        .filter(TenantGroup.company_id == resolved_company_id, TenantGroup.is_active == True)
        .order_by(TenantGroup.name.asc())
        .all()
    )
    for row_idx, group in enumerate(groups, 2):
        groups_ws.cell(row=row_idx, column=1, value=group.group_id)
        groups_ws.cell(row=row_idx, column=2, value=group.name)
        groups_ws.cell(row=row_idx, column=3, value=group.code)
        groups_ws.cell(row=row_idx, column=4, value=group.is_default)

    # Auto-width
    for sheet in (ws, groups_ws):
        for col in sheet.columns:
            max_len = max(len(str(c.value or "")) for c in col) + 2
            sheet.column_dimensions[col[0].column_letter].width = max_len

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tenant_import_template.xlsx"},
    )


@router.post("/import")
def import_tenants(
    file: UploadFile = File(...),
    group_id: int | None = Query(default=None, ge=1),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """
    Bulk-create tenants from an uploaded Excel file.

    The file must follow the template from `GET /tenants/template`.
    Returns a summary with created count and per-row errors.
    """
    _require_tenant_manager(current_user)
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File must be .xlsx or .xls")

    try:
        wb = load_workbook(filename=BytesIO(file.file.read()), read_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Could not read Excel file")

    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [
        str(cell or "").strip().rstrip("*").lower()
        for cell in header_row
    ]
    header_index = {header: idx for idx, header in enumerate(headers) if header}
    if "full_name" not in header_index:
        header_index = {
            "full_name": 0,
            "email": 1,
            "phone": 2,
            "tenant_type": 3,
        }
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file has no data rows")

    default_company_id = _resolve_company_id(None, current_user)
    if group_id is not None:
        validate_group_selection(default_company_id, group_id, db)
    created: list[dict] = []
    errors: list[dict] = []

    def cell_value(row: tuple, name: str):
        index = header_index.get(name)
        if index is None or index >= len(row):
            return None
        value = row[index]
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            return value or None
        return value

    def parse_bool(value, default: bool = True) -> bool:
        if value is None:
            return default
        text = str(value).strip().lower()
        if text in ("1", "true", "yes", "y", "active"):
            return True
        if text in ("0", "false", "no", "n", "inactive"):
            return False
        return default

    for row_num, row in enumerate(rows, start=2):
        # Skip completely empty rows
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        full_name = str(cell_value(row, "full_name") or "").strip()
        if not full_name:
            errors.append({"row": row_num, "error": "full_name is required"})
            continue
        if len(full_name) > 15:
            errors.append({"row": row_num, "error": f"full_name '{full_name}' exceeds 15 chars (device limit)"})
            continue

        try:
            row_company_id = cell_value(row, "company_id")
            target_company_id = _resolve_company_id(UUID(str(row_company_id)) if row_company_id else None, current_user)
            if target_company_id is None:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="company_id is required")

            row_group_id = cell_value(row, "group_id")
            target_group_id = int(row_group_id) if row_group_id not in (None, "") else group_id
            if target_group_id is None:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="group_id is required")
            validate_group_selection(target_company_id, target_group_id, db)

            payload = TenantCreate(
                company_id=target_company_id if current_user.role == UserRole.super_admin.value else None,
                external_id=cell_value(row, "external_id"),
                full_name=full_name,
                email=cell_value(row, "email"),
                phone=cell_value(row, "phone"),
                tenant_type=cell_value(row, "tenant_type") or "employee",
                is_active=parse_bool(cell_value(row, "is_active"), default=True),
                global_access_from=cell_value(row, "global_access_from"),
                global_access_till=cell_value(row, "global_access_till"),
                group_id=target_group_id,
            )
            tenant = create_tenant(payload, target_company_id, db)
            created.append({"row": row_num, "tenant_id": tenant.tenant_id, "full_name": tenant.full_name})
        except HTTPException as exc:
            errors.append({"row": row_num, "error": exc.detail})
        except ValueError as exc:
            errors.append({"row": row_num, "error": str(exc)})
        except Exception as exc:
            errors.append({"row": row_num, "error": str(exc)})

    return {
        "total_rows": len([r for r in rows if r and any(cell is not None and str(cell).strip() != "" for cell in r)]),
        "created": len(created),
        "failed": len(errors),
        "created_tenants": created,
        "errors": errors,
    }


@router.get("/{tenant_id}", response_model=TenantRead)
def get_tenant_route(
    tenant_id: int,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> TenantRead:
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return _to_tenant_read(tenant, db)


@router.patch("/{tenant_id}", response_model=TenantRead)
def update_tenant_route(
    tenant_id: int,
    payload: TenantUpdate,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> TenantRead:
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    tenant = update_tenant(tenant_id, payload, db)
    return _to_tenant_read(tenant, db)


@router.delete("/{tenant_id}")
def delete_tenant_route(
    tenant_id: int,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict[str, str]:
    """Delete tenant and remove their related data."""
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)

    delete_tenant_with_related_data(tenant_id, db, performed_by=current_user.user_id)
    return {"message": "Tenant deleted"}


# ---------------------------------------------------------------------------
# Fingerprint capture  (Page 2 — Enrollment)
# ---------------------------------------------------------------------------


@router.post("/{tenant_id}/capture-fingerprint")
def capture_fingerprint_route(
    tenant_id: int,
    payload: CaptureRequest,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """
    **Primary enrollment endpoint — Page 2.**

    Creates the user on the device, triggers fingerprint enrollment mode,
    polls for up to `capture_wait_seconds` for the user to scan, then
    extracts and stores the fingerprint template in the DB.

    For push-mode devices: returns immediately with a correlation_id.
    For direct-mode devices: blocks for up to capture_wait_seconds.

    Supply `valid_from` / `valid_till` to set a per-device access window that
    overrides the tenant's global dates for this specific device.
    """
    _log.info(">>> capture-fingerprint called: tenant=%d device=%d", tenant_id, payload.device_id)
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return register_and_capture_fingerprint(
        tenant_id=tenant_id,
        device_id=payload.device_id,
        db=db,
        finger_index=payload.finger_index,
        performed_by=current_user.user_id,
        valid_from=payload.valid_from,
        valid_till=payload.valid_till,
    )


@router.post("/{tenant_id}/extract-fingerprint")
def extract_fingerprint_route(
    tenant_id: int,
    device_id: int = Body(..., embed=True, description="Device the user has already scanned their finger on"),
    finger_index: int = Body(default=1, ge=1, le=10, embed=True),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """
    **Fallback** — pull a fingerprint template that the user has already
    enrolled at the device and store it in the DB.

    Use this when `capture-fingerprint` timed out but the user scanned
    their finger at the device afterwards.
    """
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return extract_fingerprint_from_device(
        tenant_id=tenant_id,
        device_id=device_id,
        db=db,
        finger_index=finger_index,
        performed_by=current_user.user_id,
    )


# ---------------------------------------------------------------------------
# Push fingerprint to other devices
# ---------------------------------------------------------------------------


@router.post("/{tenant_id}/enroll")
def enroll_route(
    tenant_id: int,
    payload: DeviceEnrollRequest,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """
    Push tenant + stored fingerprint to a device.
    No physical presence needed — uses the template from capture-fingerprint.

    Supply `valid_from` / `valid_till` to set a per-device access window that
    overrides the tenant's global dates for this specific device.

    NOTE: If tenant has no stored fingerprint, this only creates the user
    on the device. Use capture-fingerprint first for new tenants.
    """
    _log.info(">>> enroll called: tenant=%d device=%d", tenant_id, payload.device_id)
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return enroll_to_device(
        tenant_id=tenant_id,
        device_id=payload.device_id,
        db=db,
        finger_index=payload.finger_index,
        performed_by=current_user.user_id,
        valid_from=payload.valid_from,
        valid_till=payload.valid_till,
    )


@router.post("/{tenant_id}/enroll-bulk")
def enroll_bulk_route(
    tenant_id: int,
    payload: BulkEnrollRequest,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Push tenant + stored fingerprint to multiple devices at once.

    Each device entry can carry its own `valid_from` / `valid_till` so you can
    set different access windows per device in a single request.

    Example body:
    ```json
    {
      "finger_index": 1,
      "devices": [
        {"device_id": 8, "valid_till": "2026-12-31T23:59:59Z"},
        {"device_id": 9, "valid_from": "2026-06-01T00:00:00Z", "valid_till": "2026-09-30T23:59:59Z"},
        {"device_id": 10}
      ]
    }
    ```
    """
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return enroll_to_devices_bulk(
        tenant_id=tenant_id,
        devices=[item.model_dump() for item in payload.devices],
        db=db,
        finger_index=payload.finger_index,
        performed_by=current_user.user_id,
    )


@router.post("/{tenant_id}/enroll-site")
def enroll_site_route(
    tenant_id: int,
    payload: SiteEnrollRequest,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Grant a tenant access to a site and enroll them on every active device in that site.

    Single endpoint that does everything:
      - Records TenantSiteAccess in DB (the permission record)
      - Records TenantDeviceAccess per device (links device to site access)
      - Queues push enrollment commands for every active device in the site
        (create user + push fingerprint template)

    Returns a correlation_id per device. Poll each via
    GET /push/operations/{correlation_id} to track completion.
    """
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return enroll_to_site(
        tenant_id=tenant_id,
        site_id=payload.site_id,
        db=db,
        finger_index=payload.finger_index,
        valid_from=payload.valid_from,
        valid_till=payload.valid_till,
        performed_by=current_user.user_id,
    )


# ---------------------------------------------------------------------------
# Sync & unenroll
# ---------------------------------------------------------------------------


@router.put("/{tenant_id}/sync-device")
def sync_device_route(
    tenant_id: int,
    device_id: int = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Re-sync tenant details + fingerprint on a single device."""
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return update_tenant_on_device(tenant_id=tenant_id, device_id=device_id, db=db, performed_by=current_user.user_id)


@router.put("/{tenant_id}/sync-devices")
def sync_devices_bulk_route(
    tenant_id: int,
    device_ids: list[int] = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Re-sync tenant details + fingerprint on multiple devices."""
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return update_tenant_on_devices_bulk(tenant_id=tenant_id, device_ids=device_ids, db=db, performed_by=current_user.user_id)


@router.delete("/{tenant_id}/unenroll")
def unenroll_route(
    tenant_id: int,
    device_id: int = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Remove tenant from a single device (deletes user + fingerprint on device)."""
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return unenroll_from_device(tenant_id=tenant_id, device_id=device_id, db=db, performed_by=current_user.user_id)


@router.delete("/{tenant_id}/unenroll-bulk")
def unenroll_bulk_route(
    tenant_id: int,
    device_ids: list[int] = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Remove tenant from multiple devices."""
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return unenroll_from_devices_bulk(tenant_id=tenant_id, device_ids=device_ids, db=db, performed_by=current_user.user_id)


# ---------------------------------------------------------------------------
# Push enrollment status
# ---------------------------------------------------------------------------


@router.get("/{tenant_id}/enrollment-status/{correlation_id}")
def enrollment_status_route(
    tenant_id: int,
    correlation_id: str,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Check the status of an async push-mode enrollment operation.

    Convenience wrapper around GET /api/push/operations/{correlation_id}.
    Returns the aggregate status of all commands/configs queued for this enrollment.
    """
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)

    from app.api.services.push.route import get_operation_status_for_user
    return get_operation_status_for_user(correlation_id=correlation_id, db=db, current_user=current_user)


# ---------------------------------------------------------------------------
# Per-device access management
# ---------------------------------------------------------------------------


@router.get("/{tenant_id}/device-access", response_model=list[DeviceAccessRead])
def list_device_access_route(
    tenant_id: int,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> list[DeviceAccessRead]:
    """Return all enrolled devices for a tenant with their per-device validity windows.

    Use this to read back the current `valid_from` / `valid_till` stored per device,
    so the frontend can show what access window each device has.
    """
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)

    mappings = (
        db.query(DeviceUserMapping)
        .filter(DeviceUserMapping.tenant_id == tenant_id)
        .order_by(DeviceUserMapping.device_id)
        .all()
    )
    return [DeviceAccessRead.model_validate(m) for m in mappings]


@router.patch("/{tenant_id}/device-access/{device_id}")
def update_device_access_route(
    tenant_id: int,
    device_id: int,
    payload: DeviceAccessUpdate,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Update the validity window for a tenant on a specific device.

    This is the lightweight alternative to a full re-enroll — it only updates
    the dates and re-syncs the user config on the device (no fingerprint re-scan
    needed).

    Set `valid_till` to a future date to extend access, or to `null` to clear
    the per-device override and fall back to the tenant's global dates.
    """
    _require_tenant_manager(current_user)
    tenant = get_tenant(tenant_id, db)
    _check_tenant_access(tenant, current_user)
    return update_device_access_validity(
        tenant_id=tenant_id,
        device_id=device_id,
        db=db,
        valid_from=payload.valid_from,
        valid_till=payload.valid_till,
        performed_by=current_user.user_id,
    )


# ---------------------------------------------------------------------------
# Admin — device-level operations
# ---------------------------------------------------------------------------


@router.delete("/devices/{device_id}/users")
def wipe_device_users(
    device_id: int,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Queue DELETE_USER commands for all enrolled users on a device and clear DB mappings. Admin only."""
    if current_user.role not in (UserRole.super_admin.value, UserRole.company_admin.value):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")

    device = _get_device_or_404(device_id, db)
    _check_device_company_access(device, current_user)

    from app.api.services.push.commands import push_delete_credential, push_delete_user
    mappings = db.query(DeviceUserMapping).filter(DeviceUserMapping.device_id == device_id).all()
    queued = 0
    for mapping in mappings:
        push_delete_credential(db, device_id, mapping.tenant_id, cred_type="1")
        push_delete_user(db, device_id, mapping.tenant_id)
        db.delete(mapping)
        queued += 1

    db.commit()

    return {
        "device_id": device_id,
        "queued": queued,
        "message": f"DELETE_CREDENTIAL + DELETE_USER queued for {queued} user(s). Device will process on next poll.",
    }


@router.post("/devices/{device_id}/cleanup-orphans")
def cleanup_device_orphans(
    device_id: int,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
) -> dict:
    """Remove DB mappings for users no longer active. Queues DELETE_USER on the device for each."""
    if current_user.role not in (UserRole.super_admin.value, UserRole.company_admin.value):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")

    device = _get_device_or_404(device_id, db)
    _check_device_company_access(device, current_user)

    from app.api.services.push.commands import push_delete_credential, push_delete_user
    from database.models import Tenant

    # Find mappings where the tenant is inactive
    mappings = (
        db.query(DeviceUserMapping)
        .join(Tenant, Tenant.tenant_id == DeviceUserMapping.tenant_id)
        .filter(DeviceUserMapping.device_id == device_id, Tenant.is_active == False)
        .all()
    )

    queued = 0
    for mapping in mappings:
        push_delete_credential(db, device_id, mapping.tenant_id, cred_type="1")
        push_delete_user(db, device_id, mapping.tenant_id)
        db.delete(mapping)
        queued += 1

    db.commit()

    return {
        "device_id": device_id,
        "queued": queued,
        "message": f"Queued removal of {queued} inactive user(s). Device will process on next poll.",
    }
